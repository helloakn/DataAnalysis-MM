#!/usr/bin/python
import os,sys #to execute shell command
from openpyxl import load_workbook #to read xlsx file
import pyexcel as p #to convert from xls to xlsx
import xlsxwriter #to write excel file
import csv #to generate csv
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

class GeneratorLen(object):
    def __init__(self, gen, length):
        self.gen = gen
        self.length = length

    def __len__(self): 
        return self.length

    def __iter__(self):
        return self.gen
class program(object):
    #Variables
    exCurrentLine = 1
    _Path = os.path.dirname(os.path.realpath(__file__))
    _Data =  _Path +"/Data/"; _OutPut = "OutPut"; _Rule="Rule";
    _QCNameList = _Data + "/QC Name.xlsx"
    _QCName1 = None
    _file_to_check = None; _qc1=None; _qc2=None;
    _FILENAME = None
    _Inherience_Word = None
    _Rude_Word = None; _Not_Rude_Word=None
    _SpecialCharacter=None
    _SingleCharager=None
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook(_Path+'/Output/demo.xlsx')
    worksheet = workbook.add_worksheet()
    # Widen the first column to make the text clearer.
    worksheet.set_column('A:A', 20)
    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    # Write some simple text.
    worksheet.write('A1', 'Name', bold)
    worksheet.write('B1', 'File', bold)
    worksheet.write('C1', 'Line', bold)
    worksheet.write('D1', 'Original Text', bold)
    worksheet.write('E1', 'Status', bold)

    
    def writeLine(self,linenumber,text,status):
        li = self.exCurrentLine = self.exCurrentLine +1
        self.worksheet.write('A'+str(li), self._QCName1)
        self.worksheet.write('B'+str(li), self._FILENAME)
        self.worksheet.write('C'+str(li), linenumber)
        self.worksheet.write('D'+str(li), text.decode('utf-8'))
        self.worksheet.write('E'+str(li), status)

    def run(self):
        self.load_qcname_listing()

    def load_qcname_listing(self):
        wb = load_workbook(self._QCNameList)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for x in range(3, 83):
            self._file_to_check = sheet['B'+str(x)].value
            self._QCName1 = sheet['C'+str(x)].value
            print bcolors.OKGREEN+"Scaning this file : "+bcolors.ENDC+self._file_to_check
            _file = self.check_file_exists(self._file_to_check)
            if _file is not None:
                self.load_raw_data(_file)
            else:
                print bcolors.WARNING + "There is no file!"+bcolors.ENDC
        
        self.workbook.close()

    def load_raw_data(self,_file):
        #print _file
        #print self._file_to_check
        #exit()
        with open("OutPut/CSV/"+self._file_to_check+".csv", "wb") as csv_file:
            writer = csv.writer(csv_file, delimiter=',')
            wb = load_workbook(_file)
            t_autiofile = None; t_text = None;t_status = None

            sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            for x in range(1, 1001):
                print "---------------------Start----------------------"
                t_autiofile =sheet['A'+str(x)].value

                print bcolors.OKBLUE+ "Line At "+bcolors.ENDC + 'B'+str(x)
                if sheet['C'+str(x)].value is not None:
                    print bcolors.WARNING+ sheet['C'+str(x)].value.encode('utf8') +bcolors.ENDC
                if sheet['B'+str(x)].value is not None:
                    print sheet['B'+str(x)].value.encode('utf8') 

                if sheet['B'+str(x)].value is None:
                    t_text = ""
                    print bcolors.FAIL + "Null Value " + bcolors.ENDC
                    if sheet['C'+str(x)].value.replace(" ","") !="NG":
                        t_text = "#NG"
                        t_status = "NG"
                        self.writeLine('B'+str(x),"null value","Must BE NG")
                elif sheet['B'+str(x)].value.replace(" ","") == "":
                    if sheet['C'+str(x)].value != "NG":
                        print bcolors.FAIL + "Null Value " + bcolors.ENDC
                        t_text = "#NG"
                        t_status = "NG"
                        self.writeLine('B'+str(x),"null value","Must BE NG")
                else:
                    #null values passed
                    text = sheet['B'+str(x)].value.encode('utf8')
                    if self.check_Rude_Word(text):
                        #print "rube word detct"
                        if sheet['C'+str(x)].value != 'IE':
                            #it must be IE
                            self.writeLine('B'+str(x),text,"Must BE IE")
                            t_text = text
                            t_status = "IE"
                    elif self.check_NG_Word(text,sheet['C'+str(x)].value):
                            self.writeLine('B'+str(x),text,"Must BE NG  ")
                            t_text = text
                            t_status = "NG"
                    elif self.check_Spelling(text):
                            if sheet['C'+str(x)].value != 'NG':
                                self.writeLine('B'+str(x),text,"Wrong Spelling")
                                t_status = "NG"
                                print bcolors.FAIL + "Wrong Spelling " + bcolors.ENDC
                            else:
                                t_text = text
                                t_status = sheet['C'+str(x)].value
                    else:
                        t_text = text
                        t_status = sheet['C'+str(x)].value
                        print "pass"

                    #check the inherience word and replace 
                    print "Checking Inherience Word"   
                    #text  ="asdf asdf Container asdf awe"
                    #check english word 
                    if t_text is None: t_text=""
                    
                    et = (k for k, v in program._Inherience_Word.iteritems() if k in t_text) 
                    eh = GeneratorLen(et, 0)
                    eh1 = list(eh) 
                    if len(eh1)!=0:
                        etx = "["+eh1[0] +"/"+ program._Inherience_Word[eh1[0]]+"]"
                        #print etx
                        if etx not in t_text:
                            t_text = t_text.replace(eh1[0],etx)
                    #check english word 
                    mt = (k for k, v in program._Inherience_Word.iteritems() if v in t_text) 
                    mh = GeneratorLen(et, 0)
                    mh1 = list(mh) 
                    if len(mh1)!=0:
                        mtx = "["+mh1[0] +"/"+ program._Inherience_Word[mh1[0]]+"]"
                        #print etx
                        if mtx not in t_text:
                            t_text = t_text.replace(mh1[0],mtx)        
                            
                        #print t_text    
                        #exit();
                    t_autiofile = None; t_text = None;t_status = None 
                       
                    data = [t_autiofile,t_text,t_status,'','']
                    writer.writerow(data)
                    print sheet['A'+str(x)].value
                print "---------------------End------------------------\n"
                    
    def check_file_exists(self,filename):
        self._FILENAME = filename
        listOfDirs = [self._Data+f for f in os.listdir(self._Data) if os.path.isdir(self._Data+f)]
        for dir in listOfDirs:
            if os.path.exists(dir+"/"+filename+".xls"):
                print bcolors.OKGREEN+ "Converting file format "+bcolors.ENDC+"from "+bcolors.OKBLUE+"xls"+bcolors.ENDC+" to "+bcolors.OKBLUE+"xlsx"+bcolors.ENDC
                p.save_book_as(file_name=dir+"/"+filename+".xls",
                dest_file_name=self._Path+"/Output/tmp/"+filename+".xlsx")
                return self._Path+"/Output/tmp/"+filename+".xlsx"
            elif os.path.exists(dir+"/"+filename+".xlsx"):
                return dir+"/"+filename+".xlsx"
        return None
    
    def check_NG_Word(self,text,stat):
        print "checking NG word"
        if text == "NG" or text == "#NG":
            if stat == "NG":
                return False
            else:
                return True
        else:
            return False

    def check_Spelling (self,text):
        print "checking Spelling"
        t = (x for x in self._SpecialCharacter if (" "+x) in text) 
        h = GeneratorLen(t, 0)
        if (len(list(h))==0):
            return False
            t = (x for x in self._SpecialCharacter if (x+x+x) in text) 
            h = GeneratorLen(t, 0)
            if (len(list(h))==0):
                return False
            else:
                return True
        else:
            #print bcolors.FAIL+"Wrong Spelling!"+bcolors.ENDC
            return True
  
    def check_Rude_Word(self,text):
        print "checking Rube word"
        t = (x for x in self._Rude_Word if x in text) 
        h = GeneratorLen(t, 0)
        h = list(h) 
        if (len(h)==0):
            return False
        else:
            if len(self._Not_Rude_Word[h[0]])!=0:
                t = (x for x in self._Not_Rude_Word[h[0]] if x in text) 
                h = GeneratorLen(t, 0)
                h = list(h) 
                if (len(h)==0):
                    data = ["first_name,last_name,city".split(","),
                            "Tyrese,Hirthe,Strackeport".split(","),
                            "Jules,Dicki,Lake Nickolasville".split(","),
                            "Dedric,Medhurst,Stiedemannberg".split(",")
                            ]
                    path = "output.csv"
                    self.csv_writer(data, path)
                    return True
                else:
                    #print bcolors.FAIL+"Detected Rude word!"+bcolors.ENDC
                    #print text
                    #exit()
                    return False
            else:
                print bcolors.FAIL+"Detected Rude word!"+bcolors.ENDC
                return True    
     
    def check_Inherience_Word(self,text):
        print "checking check_Inherience_Word"

    def check_Bad_Word():
        print "checking check_Inherience_Word"

    def csv_writer(self,data, path):
        """
        Write data to a CSV file path
        """
        with open(path, "wb") as csv_file:
            writer = csv.writer(csv_file, delimiter=',')
            for line in data:
                writer.writerow(line)    

if __name__ == '__main__':
    print "============================================================="
    print "=================== Developed by ACE-PLus ==================="
    print "============================================================="
    print "Initializating..."
    #Inherience Words
    print bcolors.OKGREEN+ "Loading Inherience Words..."+bcolors.ENDC
    program._Inherience_Word = {}
    _file = program._Path+"/Rule/Inherience-Word.xlsx"
    wb = load_workbook(_file)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    for row in sheet:
        if(row[0].value is not None):
            eng= row[0].value.encode('utf8')
            mm= row[2].value.encode('utf8')
            program._Inherience_Word[eng] = mm
    #Rude Words
    print bcolors.OKGREEN+ "Loading Rude Words..."+bcolors.ENDC
    program._Rude_Word = []
    program._Not_Rude_Word = {}
    _file = program._Path+"/Rule/Rude-Word.xlsx"
    wb = load_workbook(_file)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    for row in sheet:
        if(row[0].value is not None):
            word= row[0].value.encode('utf8')
            program._Rude_Word.append(word)
            if(row[1].value is not None):
                program._Not_Rude_Word[word] =row[1].value.encode('utf8').split(',')
            else:
                program._Not_Rude_Word[word] ={}
            
    #character
    print bcolors.OKGREEN+ "Loading Characters..."+bcolors.ENDC
    program._SpecialCharacter = []
    program._SingleCharager = []
    _file = program._Path+"/Rule/Special-Character.xlsx"
    wb = load_workbook(_file)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    for row in sheet:
        if(row[0].value is not None):
            word= row[0].value.encode('utf8')
            program._SpecialCharacter.append(word)
        if(row[1].value is not None):
            word= row[1].value.encode('utf8')
            program._SingleCharager.append(word)
    #print program._SpecialCharacter[1]
    #print program._Not_Rude_Word
    search_age="Container"
    #for name, age in program._Inherience_Word.iteritems():
    #    if name == search_age:
    #        print name
    #text  ="asdf asdf Container asdf awe"
    #t = (k for k, v in program._Inherience_Word.iteritems() if k in text) 
    #h = GeneratorLen(t, 0)
    #h = list(h) 
    #print h
    program().run()